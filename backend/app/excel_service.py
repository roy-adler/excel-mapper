from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from sqlalchemy.orm import Session

from app.config import settings
from app.models import Session as MapperSession
from app.schemas import MappingRule


def ensure_storage_dirs() -> None:
    base = Path(settings.storage_dir)
    (base / "templates").mkdir(parents=True, exist_ok=True)
    (base / "sessions").mkdir(parents=True, exist_ok=True)


def expand_range(cell_range: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cells = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append(f"{get_column_letter(col)}{row}")
    return cells


def normalize_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def cast_value(value: Any, expected_type: str):
    if value in ("", None):
        return None
    if expected_type == "string":
        return str(value)
    if expected_type == "number":
        try:
            return float(value)
        except (TypeError, ValueError):
            raise ValueError("must be a number")
    if expected_type == "boolean":
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in {"true", "1", "yes", "y"}:
                return True
            if lowered in {"false", "0", "no", "n"}:
                return False
        raise ValueError("must be a boolean")
    if expected_type == "date":
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError("must be a date in YYYY-MM-DD format")
        raise ValueError("must be a date")
    raise ValueError("unsupported type")


def build_fields_from_workbook(workbook_path: str, schema_json: list[dict[str, Any]]) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path)
    fields: list[dict[str, Any]] = []
    try:
        for rule_dict in schema_json:
            rule = MappingRule(**rule_dict)
            if rule.sheet not in workbook.sheetnames:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Sheet '{rule.sheet}' not found",
                )
            ws = workbook[rule.sheet]
            for cell in expand_range(rule.range):
                value = ws[cell].value
                display_value = normalize_date(value) if rule.type == "date" else value
                fields.append(
                    {
                        "key": f"{rule.sheet}!{cell}",
                        "sheet": rule.sheet,
                        "cell": cell,
                        "type": rule.type,
                        "label": rule.label or f"{rule.sheet}:{cell}",
                        "value": display_value,
                    }
                )
    finally:
        workbook.close()
    return fields


def update_workbook_values(
    workbook_path: str,
    schema_json: list[dict[str, Any]],
    values: dict[str, Any],
) -> None:
    workbook = load_workbook(workbook_path)
    try:
        allowed_keys: dict[str, str] = {}
        for rule_dict in schema_json:
            rule = MappingRule(**rule_dict)
            ws = workbook[rule.sheet]
            for cell in expand_range(rule.range):
                allowed_keys[f"{rule.sheet}!{cell}"] = rule.type
                if f"{rule.sheet}!{cell}" in values:
                    try:
                        ws[cell].value = cast_value(values[f"{rule.sheet}!{cell}"], rule.type)
                    except ValueError as exc:
                        raise HTTPException(
                            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail=f"Invalid value for {rule.sheet}!{cell}: {exc}",
                        )
        unknown = set(values.keys()) - set(allowed_keys.keys())
        if unknown:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown mapped fields: {sorted(unknown)}",
            )
        workbook.save(workbook_path)
    finally:
        workbook.close()


def touch_session(db: Session, session_obj: MapperSession) -> None:
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    session_obj.last_activity_at = now
    session_obj.expires_at = now + timedelta(hours=settings.session_ttl_hours)
    db.add(session_obj)
    db.commit()
    db.refresh(session_obj)


def ensure_session_active(session_obj: MapperSession) -> None:
    if session_obj.expires_at < datetime.utcnow():
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="Session expired")
